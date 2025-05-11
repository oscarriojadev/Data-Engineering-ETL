# excel_report_generator.py

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

def create_excel_file(file_path):
    """
    Create a new Excel file.

    Parameters:
    file_path (str): Path to the Excel file.

    Returns:
    openpyxl.Workbook: Excel workbook.
    """
    workbook = openpyxl.Workbook()
    workbook.save(file_path)
    return workbook

def add_sheet(workbook, sheet_name):
    """
    Add a new sheet to the Excel workbook.

    Parameters:
    workbook (openpyxl.Workbook): Excel workbook.
    sheet_name (str): Name of the sheet.

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Excel worksheet.
    """
    workbook.create_sheet(sheet_name)
    return workbook[sheet_name]

def write_data_to_sheet(sheet, data, start_row=1, start_col=1):
    """
    Write data to the Excel sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet.
    data (pd.DataFrame): Data to write.
    start_row (int): Starting row.
    start_col (int): Starting column.
    """
    for row_idx, row in enumerate(data.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row, start=start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)

def format_header(sheet, header_row, header_font=Font(bold=True), header_alignment=Alignment(horizontal='center')):
    """
    Format the header row in the Excel sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet.
    header_row (int): Row number of the header.
    header_font (openpyxl.styles.Font): Font style for the header.
    header_alignment (openpyxl.styles.Alignment): Alignment for the header.
    """
    for cell in sheet[header_row]:
        cell.font = header_font
        cell.alignment = header_alignment

def format_data(sheet, start_row, end_row, start_col, end_col, data_font=Font(), data_alignment=Alignment(horizontal='left')):
    """
    Format the data cells in the Excel sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet.
    start_row (int): Starting row.
    end_row (int): Ending row.
    start_col (int): Starting column.
    end_col (int): Ending column.
    data_font (openpyxl.styles.Font): Font style for the data.
    data_alignment (openpyxl.styles.Alignment): Alignment for the data.
    """
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment

def add_border(sheet, start_row, end_row, start_col, end_col, border_style=Side(border_style='thin')):
    """
    Add borders to the cells in the Excel sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet.
    start_row (int): Starting row.
    end_row (int): Ending row.
    start_col (int): Starting column.
    end_col (int): Ending column.
    border_style (openpyxl.styles.Side): Border style.
    """
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border

def add_chart(sheet, chart_type, data_range, chart_title, chart_position):
    """
    Add a chart to the Excel sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet.
    chart_type (str): Type of chart ('bar', 'line', 'pie').
    data_range (str): Range of data for the chart.
    chart_title (str): Title of the chart.
    chart_position (str): Position of the chart (e.g., 'A10').
    """
    if chart_type == 'bar':
        chart = BarChart()
    # Add more chart types as needed

    chart.title = chart_title
    chart.set_categories(Reference(sheet, min_col=data_range.split(':')[0][0], min_row=data_range.split(':')[0][1:], max_row=data_range.split(':')[1][1:]))
    chart.add_data(Reference(sheet, min_col=data_range.split(':')[1][0], min_row=data_range.split(':')[0][1:], max_row=data_range.split(':')[1][1:]))

    sheet.add_chart(chart, chart_position)

def main():
    # Example usage
    file_path = 'report.xlsx'
    data = pd.DataFrame({
        'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35],
        'Salary': [50000, 60000, 70000]
    })

    workbook = create_excel_file(file_path)
    sheet = add_sheet(workbook, 'Report')

    write_data_to_sheet(sheet, data)
    format_header(sheet, header_row=1)
    format_data(sheet, start_row=2, end_row=len(data) + 1, start_col=1, end_col=len(data.columns))
    add_border(sheet, start_row=1, end_row=len(data) + 1, start_col=1, end_col=len(data.columns))

    add_chart(sheet, chart_type='bar', data_range='A1:C4', chart_title='Salary by Age', chart_position='A10')

    workbook.save(file_path)
    print(f'Excel report generated at {file_path}')

if __name__ == '__main__':
    main()
